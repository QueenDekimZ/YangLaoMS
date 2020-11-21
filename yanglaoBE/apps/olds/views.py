from django.shortcuts import render
# 导入Olds类
from olds.models import Olds
# 导入JsonResponse模块
from django.http import JsonResponse
import json
# 导入Q查询
from django.db.models import Q
from django.conf import settings

import openpyxl
import os
import uuid
import hashlib

def login(request):
    data_login = request.body.decode('utf-8')
    data = json.loads(data_login)
    print(data)
    if(data['username'] == 'admin' and data['password'] == '123456'):
        return JsonResponse({'code':1})
    else:
        return JsonResponse({'code':0,'msg':'账户名或密码输入错误！'})

def get_olds(request):
    """获取所有老人的信息"""
    try:
        # 使用ORM获取所有老人的信息,并把对象转为字典格式
        obj_olds = Olds.objects.all().values()
        # 把结果转为List
        olds = list(obj_olds)
        # 返回

        return JsonResponse({'code':1,'data':olds})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code':0,'msg':"加载老人信息失败！异常为" + str(e)})

def query_olds(request):
    """查询老人信息"""
    try:
        # 接收短笛过来的查询条件  ---axios默认是json ---字典类型('inputstr') data['inputstr']
        data1 = request.body.decode('utf-8')
        data = json.loads(data1)
        # 使用ORM获取所有满足条件的老人的信息
        obj_olds = Olds.objects.filter(Q(oNo__icontains=data['inputstr'])|
                                       Q(oName__icontains=data['inputstr'])|
                                       Q(oIdentity__icontains=data['inputstr'])|
                                       Q(oGender__icontains=data['inputstr'])|
                                       Q(oBirthday__icontains=data['inputstr'])|
                                       Q(oMobile__icontains=data['inputstr'])|
                                       Q(oFamilyMobile__icontains=data['inputstr'])|
                                       Q(oAddress__icontains=data['inputstr'])|
                                       Q(oHealth__icontains=data['inputstr'])|
                                       Q(oSocial__icontains=data['inputstr'])).values()
        # 把外层容器转为List
        olds = list(obj_olds)
        # 返回
        return JsonResponse({'code':1,'data':olds})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code':0,'msg':"查询老人信息失败！异常为" + str(e)})

def is_exists_oNo(request):
    """判断学号是否存在"""
    # 接受传递过来的学号
    data = json.loads(request.body.decode('utf-8'))
    try:
        obj_olds = Olds.objects.filter(oNo=data['oNo'])
        if obj_olds.count() == 0:
            return JsonResponse({'code':1, 'exists':False})
        else:
            return JsonResponse({'code': 1, 'exists':True})
    except Exception as e:
        return JsonResponse({'code':0,'msg':"校验编号失败！异常为" + str(e)})

def add_olds(request):
    """添加老人信息到数据库"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode('utf-8'))
    print(data)
    try:
        # 添加到数据库
        obj_olds = Olds(oNo=data['oNo'],oIdentity=data['oIdentity'],oName=data['oName'],oGender=data['oGender'],
                        oBirthday=data['oBirthday'],oMobile=data['oMobile'],oFamilyMobile=data['oFamilyMobile'],
                        oAddress=data['oAddress'],oHealth=data['oHealth'],oSocial=data['oSocial'],oImage=data['oImage'])
        # 执行添加
        obj_olds.save()
        # 使用ORM获取所有老人的信息,并把对象转为字典格式
        obj_olds = Olds.objects.all().values()
        # 把结果转为List
        olds = list(obj_olds)
        # 返回
        return JsonResponse({'code':1,'data':olds})
    except Exception as e:
        return JsonResponse({'code':0,'msg':"添加老人信息失败!异常为：" + str(e)})

def update_olds(request):
    """修改老人信息到数据库"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 查找到要修改的老人信息
        obj_olds = Olds.objects.get(oNo=data['oNo'])
        # 依次修改
        obj_olds.oIdentity = data['oIdentity']
        obj_olds.oName = data['oName']
        obj_olds.oGender = data['oGender']
        obj_olds.oBirthday = data['oBirthday']
        obj_olds.oMobile = data['oMobile']
        obj_olds.oFamilyMobile = data['oFamilyMobile']
        obj_olds.oAddress = data['oAddress']
        obj_olds.oHealth = data['oHealth']
        obj_olds.oSocial = data['oSocial']
        obj_olds.oImage = data['oImage']
        # 保存
        obj_olds.save()
        # 使用ORM获取所有老人的信息,并把对象转为字典格式
        obj_olds = Olds.objects.all().values()
        # 把结果转为List
        olds = list(obj_olds)
        # 返回
        return JsonResponse({'code':1,'data':olds})
    except Exception as e:
        return JsonResponse({'code':0,'msg':"更新老人信息失败!异常为：" + str(e)})

def delete_olds(request):
    """删除数据库中老人信息"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 查找到要修改的老人信息
        obj_olds = Olds.objects.get(oNo=data['oNo'])
        # 执行删除
        obj_olds.delete()
        # 使用ORM获取所有老人的信息,并把对象转为字典格式`
        obj_olds = Olds.objects.all().values()
        # 把结果转为List
        olds = list(obj_olds)
        # 返回
        return JsonResponse({'code':1,'data':olds})
    except Exception as e:
        return JsonResponse({'code':0,'msg':"删除老人信息失败!异常为：" + str(e)})

def delete_olds_all(request):
    """批量删除数据库中老人信息"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode('utf-8'))
    try:
        for one_old in data['old']:
            # 查询当前记录
            obj_olds = Olds.objects.get(oNo=one_old['oNo'])
            obj_olds.delete()
        # 使用ORM获取所有老人的信息,并把对象转为字典格式`
        obj_olds = Olds.objects.all().values()
        # 把结果转为List
        olds = list(obj_olds)
        # 返回
        return JsonResponse({'code':1,'data':olds})
    except Exception as e:
        return JsonResponse({'code':0,'msg':"批量删除老人信息失败!异常为：" + str(e)})




def upload(request):
    """接受上传的文件"""
    rev_file = request.FILES.get('avatar')
    if not rev_file:
        return JsonResponse({'code':0,'msg':'图片不存在！'})
    # 获得一个唯一的名字: uuid+hash
    new_name = get_random_str()
    # 准备写入的URL
    # file_name = new_name + os.path.splitext(rev_file)[1]
    file_name = new_name + os.path.splitext(rev_file.name)[1]

    file_path = os.path.join(settings.MEDIA_ROOT, file_name)
    # 开始写入到本地磁盘
    try:
        f = open(file_path,'wb')
        # 文件较大，系统资源不够，多次写入
        for i in rev_file.chunks():
            f.write(i)
        f.close()
        return JsonResponse({'code':1,'name':file_name})

    except Exception as e:
        return JsonResponse({'code':0,'msg':str(e)})

def get_random_str():
    # 获取uuid的随机数
    uuid_val = uuid.uuid4()
    # 获取uuid的随机数字符串
    uuid_str = str(uuid_val).encode('utf-8')
    # 获取md5实例
    md5 = hashlib.md5()
    # 获取uuid的md5摘要
    md5.update(uuid_str)
    # 返回固定长度的字符串
    return md5.hexdigest()

def export_olds_excel(request):
    """将老人信息导出到Excel"""
    # 获取所有的老人信息
    obj_olds = Olds.objects.all().values()
    olds = list(obj_olds)
    # 准备Excel文件名
    excel_name = get_random_str() + ".xlsx"
    # 准备写入的路径
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # 写入到Excel
    write_excel_dict(olds, path)

    return JsonResponse({'code':1,'name':excel_name})


def import_olds_excel(request):
    """从Excel批量导入老人信息"""
    # ==========1、接收Excel文件存储到Media文件夹==============
    rev_file = request.FILES.get('excel')
    if not rev_file:
        return JsonResponse({'code':0,'msg':'Excel文件不存在！'})
    # 获得一个唯一的名字: uuid+hash
    new_name = get_random_str()
    # 准备写入的URL
    # file_name = new_name + os.path.splitext(rev_file)[1]
    file_name = new_name + os.path.splitext(rev_file.name)[1]

    file_path = os.path.join(settings.MEDIA_ROOT, file_name)
    # 开始写入到本地磁盘
    try:
        f = open(file_path,'wb')
        # 文件较大，系统资源不够，多次写入
        for i in rev_file.chunks():
            f.write(i)
        f.close()
    except Exception as e:
        return JsonResponse({'code':0,'msg':"读取Excel文件异常"+str(e)})
    # ================2、读取存储在Media文件夹的数据===============
    ex_olds = read_excel_dict(file_path)
    # 3、把读取的数据存储到数据库
    # 定义几个变量： success：error：errors：
    success = 0
    error = 0
    error_onos = []
    for one_old in ex_olds:
        # print(one_old)
        try:

            # obj_old = Olds(oNo=one_old['oNo'], oIdentity=one_old['oIdentity'], oName=one_old['oName'], oGender=one_old['oGender'],
            #                 oBirthday=one_old['oBirthday'], oMobile=one_old['oMobile'], oFamilyMobile=one_old['oFamilyMobile'],
            #                 oAddress=one_old['oAddress'], oHealth=one_old['oHealth'], oSocial=one_old['oSocial'],
            #                 oImage=one_old['oImage'])
            #
            # obj_old.save()
            #======明确创建，查重，防止已经有的oNo数据冲突
            obj_old = Olds.objects.create(oNo=one_old['oNo'],oIdentity=one_old['oIdentity'],
                           oName=one_old['name'],oGender=one_old['oGender'],
                           oBirthday=one_old['oBirthday'],oMobile=one_old['oMobile'],
                           oFamilyMobile=one_old['oFamilyMobile'],oAddress=one_old['oAddress'],
                           oHealth=one_old['oHealth'],oSocial=one_old['oSocial'])

            success += 1
        except Exception as e:
            error += 1
            error_onos.append(one_old['oNo'])
    # 4、返回--导入信息（成功：5，失败：1--（oNo）），所有老人信息
    obj_olds = Olds.objects.all().values()
    olds = list(obj_olds)
    return JsonResponse({'code':1,'success':success,'error':error,'errors':error_onos,'data':olds})

def read_excel_dict(path:str):
    """读取Excel数据，存储为字典----[{},{},{}]"""
    # 实例化一个workbook
    workbook = openpyxl.load_workbook(path)
    # 实例化一个sheet
    sheet = workbook['old']
    # 定义一个变量存储最终的数据--[]
    olds = []
    # 准备key
    keys = ['oNo','oIdentity','oName','oGender',
            'oBirthday','oMobile','oFamilyMobile',
            'oAddress','oImage','oHealth','oSocial']
    # 遍历
    for row in sheet.rows:
        # 定义一个临时的字典
        temp_dict = {}
        # 组合值和key
        for index,cell in enumerate(row):
            # 组合
            temp_dict[keys[index]] = cell.value
        # 附加到list中
        olds.append(temp_dict)
    return olds

def write_excel_dict(data:list,path:str):
    """把数据写入Excel"""
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'old'
    # 准备keys
    keys = data[0].keys()
    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每一个元素
        for k, v in enumerate(keys):
            sheet.cell(row=index+1,column=k+1,value=str(item[v]))
    # 写入到文件
    workbook.save(path)